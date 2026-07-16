import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


def save_excel(data, filename):

    if not data:
        print("Nincs feldolgozható adat.")
        return


    df = pd.DataFrame(data)


    # Oszlopsorrend biztosítása
    columns = [
        "Dátum",
        "Értéknap",
        "Tranzakció típusa",
        "Partner",
        "Közlemény",
        "Összeg",
        "Egyenleg"
    ]


    for col in columns:
        if col not in df.columns:
            df[col] = ""


    df = df[columns]


    # Excel mentés
    df.to_excel(
        filename,
        index=False,
        sheet_name="Kivonat"
    )


    # Formázás
    wb = load_workbook(filename)

    ws = wb["Kivonat"]


    # Fejléc formázás
    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # Oszlopszélesség automatikus beállítása
    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter


        for cell in column:

            if cell.value:

                length = len(
                    str(cell.value)
                )

                if length > max_length:
                    max_length = length


        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 3,
            50
        )


    # Összeg formázás
    for row in ws.iter_rows(
        min_row=2
    ):

        for index in [6,7]:

            cell = row[index-1]

            if isinstance(cell.value, (int,float)):

                cell.number_format = '#,##0.00'


    wb.save(filename)
