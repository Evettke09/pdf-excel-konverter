import os
from raiffeisen_parser import parse_pdf
from excel_writer import save_excel


INPUT_FOLDER = "input"
OUTPUT_FOLDER = "output"


def main():

    if not os.path.exists(INPUT_FOLDER):
        os.makedirs(INPUT_FOLDER)

    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)


    pdf_files = [
        f for f in os.listdir(INPUT_FOLDER)
        if f.lower().endswith(".pdf")
    ]


    if not pdf_files:
        print("Nincs PDF az input mappában.")
        return


    all_transactions = []


    for pdf_file in pdf_files:

        path = os.path.join(INPUT_FOLDER, pdf_file)

        print(f"Feldolgozás: {pdf_file}")

        transactions = parse_pdf(path)

        all_transactions.extend(transactions)


    save_excel(
        all_transactions,
        os.path.join(
            OUTPUT_FOLDER,
            "Raiffeisen_kivonat.xlsx"
        )
    )


    print("Kész! Az Excel elkészült.")



if __name__ == "__main__":
    main()
import pdfplumber
import re
from datetime import datetime


def clean_amount(value):
    """
    Magyar formátumú összeg átalakítása számmá.
    Példa:
    -20.000,00 -> -20000.00
    """

    if not value:
        return None

    value = value.replace(" ", "")
    value = value.replace(".", "")
    value = value.replace(",", ".")

    try:
        return float(value)

    except:
        return None



def is_date(text):
    """
    Dátum felismerése
    Példa:
    2025.02.01
    """

    pattern = r"^\d{4}\.\d{2}\.\d{2}"

    return bool(
        re.match(pattern, text)
    )



def parse_pdf(pdf_path):

    transactions = []


    with pdfplumber.open(pdf_path) as pdf:


        for page in pdf.pages:

            text = page.extract_text()


            if not text:
                continue


            lines = text.split("\n")


            current = None


            for line in lines:


                line = line.strip()


                if not line:
                    continue



                # Új tranzakció kezdete
                if is_date(line):


                    # előző mentése
                    if current:
                        transactions.append(current)



                    current = {

                        "Dátum": "",
                        "Értéknap": "",
                        "Tranzakció típusa": "",
                        "Partner": "",
                        "Közlemény": "",
                        "Összeg": None,
                        "Egyenleg": None

                    }



                    parts = line.split()


                    current["Dátum"] = parts[0]


                    # számok keresése a sorban
                    numbers = []


                    for p in parts:

                        if re.search(
                            r"-?\d+\.\d+,\d{2}",
                            p
                        ):
                            numbers.append(p)



                    if numbers:

                        current["Összeg"] = clean_amount(
                            numbers[0]
                        )


                        if len(numbers) > 1:

                            current["Egyenleg"] = clean_amount(
                                numbers[-1]
                            )



                else:

                    if current is None:
                        continue



                    # összeg folytatás felismerése
                    amounts = re.findall(
                        r"-?\d+\.\d+,\d{2}",
                        line
                    )


                    if amounts:

                        current["Összeg"] = clean_amount(
                            amounts[0]
                        )


                    else:

                        # első szövegsor partner
                        if not current["Partner"]:

                            current["Partner"] = line

                        else:

                            current["Közlemény"] += (
                                " " + line
                            )



            if current:

                transactions.append(current)



    return transactions
  import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


def save_excel(data, filename):

    if not data:
        print("Nincs feldolgozható adat.")
        return


    df = pd.DataFrame(data)


    # Oszlopsorrend biztosítása
    columns = [
        "Dátum",
        "Értéknap",
        "Tranzakció típusa",
        "Partner",
        "Közlemény",
        "Összeg",
        "Egyenleg"
    ]


    for col in columns:
        if col not in df.columns:
            df[col] = ""


    df = df[columns]


    # Excel mentés
    df.to_excel(
        filename,
        index=False,
        sheet_name="Kivonat"
    )


    # Formázás
    wb = load_workbook(filename)

    ws = wb["Kivonat"]


    # Fejléc formázás
    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # Oszlopszélesség automatikus beállítása
    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter


        for cell in column:

            if cell.value:

                length = len(
                    str(cell.value)
                )

                if length > max_length:
                    max_length = length


        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 3,
            50
        )


    # Összeg formázás
    for row in ws.iter_rows(
        min_row=2
    ):

        for index in [6,7]:

            cell = row[index-1]

            if isinstance(cell.value, (int,float)):

                cell.number_format = '#,##0.00'


    wb.save(filename)
